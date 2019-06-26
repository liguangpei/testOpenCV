# -*- coding: UTF-8 -*-
import openpyxl
#加载文件
wb = openpyxl.load_workbook("../file/test.xlsx")
#获得sheet名称
sheetNames = wb.sheetnames
print(sheetNames)
#sheetName1 = sheetNames[0]
#根据名称获取第一个sheet
#sheet1 = wb[sheetName1]
#根据索引获得第一个sheet
sheet1 = wb.worksheets[0]

L = ['张三', '李四', '王五']
#excel中单元格为B2开始，即第2列，第2行
for i in range(len(L)):
    sheet1.cell(i+2, 2).value=L[i]
#保存数据，如果提示权限错误，需要关闭打开的excel
wb.save("../file/test.xlsx")